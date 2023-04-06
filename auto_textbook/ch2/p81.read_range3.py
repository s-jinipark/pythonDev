
# iter_rows 를 이용해 범위 내의 셀 얻기

import openpyxl as excel

# 워크북 열고 시트를 가져오기
book = excel.load_workbook('write_cellname.xlsx')
sheet = book.active

# 이터레이터 얻기
it = sheet.iter_rows(min_row=2, min_col=2,
                     max_row=4, max_col=4)

# for 문과 조합해 셀의 값 얻기
for row in it:
    r = []
    for cell in row:
        r.append(cell.value)
    print(r)

print('--------------------')

