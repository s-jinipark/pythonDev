
# 반복문을 이용한 연속 데이터 채우기

import openpyxl as excel

# 워크북을 생성하고 활성화된 워크시트 가져오기
book = excel.Workbook()
sheet = book.active

# A열에 연속 데이터 채우기
for i in range(10):
    # A 열 i+1 행에 데이터 쓰기
    sheet.cell(row=(i+1), column=1, value=i)

# 파일 (워크북) 저장
book.save('write_column.xlsx')

print('--------------------')

book2 = excel.Workbook()
sheet2 = book2.active

for i in range(1,10):
    for j in range(1,10):
        #sheet2.cell(row=j, column=i, value=i*j)
        # -> 이것도 되고, 책에 나온 방식은 아래
        cell = sheet2.cell(row=j, column=i)
        cell.value = i*j

book2.save('write_9x9.xlsx')
