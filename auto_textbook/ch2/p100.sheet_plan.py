
# 고객의 구매 플랜ㄴ에 따라 시트 나누기

import openpyxl as excel

# 고객 명부 문서 열고 '명부' 시트 가져오기
book = excel.load_workbook('input/all-customer.xlsx')
sheet = book["명부"]


# 고객 명부를 확인하며 시트 나누기
for row in sheet.iter_rows(min_row=3) :
    cells = [v.value for v in row]

    if cells[0] is None : break
    # 읽어온 고객 정보를 변수에 저장
    (name, area, plan) = cells
    # 붙여 넣을 시트 이름을 정하기
    sname = plan + "플랜"
    # 해당 시트가 있는지 확인
    if sname not in book.sheetnames:
        to_sheet = book.create_sheet(title=sname)
        to_sheet.append(["이름", "주소", "플랜"])
    else:
        to_sheet = book[sname]

    # 해당 시트에 고객 정보를 추가
    to_sheet.append(cells)

# 파일 저장
book.save("output/sheet_plan.xlsx")

print('--------------------')

