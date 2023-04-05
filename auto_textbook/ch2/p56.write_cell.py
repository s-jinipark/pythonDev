
# 2-3 엑셀 데이터 쓰기
# 크게 3 가지가 있다
# 셀 주소 지정하거나, 행과 열의 번호를 지정하거나
# 셀 개체를 먼저 얻은 후 값을 쓰는 방법

import openpyxl as excel

# 워크북을 생성하고 활성화된 워크시트 가져오기
book = excel.Workbook()
sheet = book.active

# A1 에 값 설정
sheet["A1"] = "일찍 일어나는 새가 벌레를 잡는다"

# A2(row=2, column=1) 에 값 설정
sheet.cell(row=2, column=1, value="하늘은 스스로 돕는자를 돕니다.")

# A3(row=3, column=1) 에 값 설정
third_cell = sheet.cell(row=3, column=1)
third_cell.value = "낙숫물이 바위를 뚫는다"

sheet.cell(4, 1, "테스트 가 테스트 했다.")

# 워크북 저장
book.save('write_cell.xlsx')
