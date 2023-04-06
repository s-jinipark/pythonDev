
# 연도ㆍ육십갑자 대응표 만들기

import openpyxl as excel
import datetime
'''
2가지 전제
 1. 서기 4년은 갑자년이다
 2. 천간이 같은 해는 10으로 나눈 나머지가 같고
    지지가 같은 해는 12로 나눈 나머지가 같다 
'''
# 표의 내용을 리스트로 나타내기
gan = ['갑을병정무기경신임계', '甲乙丙丁戊己庚辛壬癸', '청청적적황황백백흑흑' ]
ji = ['자축인묘진사오미신유술해', '子丑寅卯辰巳午未申酉戌亥',
      ['쥐','소','호랑이','토끼','용','뱀','말','양','원숭이','닭','개','돼지']]

# 특정 연도에서 간지로 변환하는 함수를 정의
def year_to_ganji(year):
    # 나머지 연산을 통해 연도를 인덱스로 변환하기
    i = (year - 4) % 10
    j = (year - 4) % 12

    # 인덱스로 간지 정보 조회하기
    ganji = gan[0][i] + ji[0][j] + '(' + gan[1][i] + ji[1][j] + ')'

    # 인덱스로 동물 정보 조회하기
    color_animal = gan[2][i] + '색' + ji[2][j]
    return ganji, color_animal

# 새 워크북을 만들고 시트 가져오기
book = excel.Workbook()
sheet = book.active

# 헤더 설정
sheet["A1"] = "연도"
sheet["B1"] = "간지"
sheet["C1"] = "동물"

# 시작년도 올해로 설정하기
start_y = now = datetime.datetime.now().year

# 100년간의 연도ㆍ간지 정보를 시트에 채우기
for i in range(100):
    # 연도를 간지로 변환하기
    year = start_y - i
    result = year_to_ganji(year)
    gangi = result[0]
    col_ani = result[1]

    # 시트에 입력하기
    sheet.cell(row=i + 2, column=1, value=str(year) + "년")
    sheet.cell(row=i + 2, column=2, value=gangi)
    sheet.cell(row=i + 2, column=3, value=col_ani)

    print(year, '=', gangi, ",", col_ani)

# 파일 (워크북) 저장
book.save('write2_ganji.xlsx')

print('--------------------')

