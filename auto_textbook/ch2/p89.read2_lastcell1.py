
# 데이터가 몇 행 있는지 확실치 않을 때

import openpyxl as excel


#book = excel.load_workbook('input/monthly_sales.xlsx', data_only=True)
#sheet = book.active

sheet = excel.load_workbook("input/monthly_sales.xlsx").active
# -> 헐 한줄에

print((sheet.max_row, sheet.max_column))

print('--------------------')

# 실제 데이터 보다 넓은 범위에 테두리가 들어가 있다면 ?

sheet = excel.load_workbook("input/monthly_sales2.xlsx").active

print((sheet.max_row, sheet.max_column))

print('--------------------')

# max_row 를 사용하지 않고, 적당한 최하행의 값을 999로 지정해 데이터를 읽음

book = excel.load_workbook('input/monthly_sales2.xlsx', data_only=True)
sheet = book.active

# 셀 A3 부터 셀 F999 (적당히 큰 범위) 를 얻기
rows = sheet["A3":"F999"]
for row in rows :
    # 셀의 값을 리스트로 얻기
    values = [cell.value for cell in row]
    # 비어 있는 셀이면 읽기를 종료
    if values[0] is None : break
    # 리스트를 출력
    print(values)

print('--------------------')

# 앞에서 적당한 범위를 지정해 모들 데이터를 가져 왔지만
# iter_rows() 메서드를 사용해도 모든 셀을 가져올 수 있다

# 위의 sheet 사용
# iter_rows() 사용
for row in sheet.iter_rows(min_row=3):  # A3 이하의 행을 얻고자 함(헤더 빼고 실 데이터..)
    values = [cell.value for cell in row]
    if values[0] is None : break
    print(values)
