
# 생년ㆍ나이 대응표 만들기

import openpyxl as excel
import datetime

# 워크북을 생성하고 활성화된 워크시트 가져오기
book = excel.Workbook()
sheet = book.active

# 올해 연도 구하기
thisyear = datetime.datetime.now().year
print(type(thisyear))

# 1행에 해더 설정
sheet["A1"] = "출생 연도"
sheet["B1"] = "세는 나이"
sheet["C1"] = "만 나이 (생일 후)"
sheet["D1"] = "만 나이 (생일 전)"

# 셀의 너비 조정
sheet.column_dimensions['C'].width=15
sheet.column_dimensions['D'].width=15

# 생년ㆍ나이 연속 데이터 채우기
for i in range(80):
    # 설정할 값을 계산
    birth_year = thisyear - i
    korean_age = thisyear - birth_year + 1  # 세는 나이는 1살 부터 시작
    man_age = {'after_bday': korean_age-1, 'before_bday':korean_age-2}

    # 셀을 읽어 값을 설정하기
    year_cell = sheet.cell(row=i+2, column=1)   # i 는 0 부터 시작이므로 + 2 해줌
    year_cell.value = str(birth_year) + "년생"

    age_cell = sheet.cell(row=i+2, column=2)
    age_cell.value = str(korean_age) + "세"

    age_cell = sheet.cell(row=i+2, column=3)
    age_cell.value = "만 " + str(man_age['after_bday']) + "세"

    age_cell = sheet.cell(row=i+2, column=4)
    age_cell.value = "만 " + str(man_age['before_bday']) + "세"

# 예외 경우 처리
sheet["D2"] = "-"

# 파일 (워크북) 저장
book.save('write2_agelist.xlsx')

print('--------------------')

